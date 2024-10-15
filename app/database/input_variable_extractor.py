# app/database/input_variable_extractor.py

import openpyxl
import pandas as pd
import logging

logger = logging.getLogger(__name__)

def find_inputs_and_extract(doc_name, sheet_name, input_header):
    try:
        logger.debug(f"Loading workbook '{doc_name}'.")
        wb = openpyxl.load_workbook(doc_name)
        sheet = wb[sheet_name]
        logger.debug(f"Loaded sheet '{sheet_name}' from workbook '{doc_name}'.")

        column_value = None
        row_value = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == input_header:
                    column_value = cell.column
                    row_value = cell.row
                    logger.debug(f"Found input header '{input_header}' at column {column_value}, row {row_value}.")
                    break
            if column_value and row_value:
                break

        if not column_value or not row_value:
            logger.error(f"Input header '{input_header}' not found in sheet '{sheet_name}'.")
            return {}

        df = pd.read_excel(doc_name, sheet_name=sheet_name, header=row_value-1)
        wanted_columns = df.iloc[:, [column_value - 1, column_value, column_value + 1]]
        df_cleaned = wanted_columns.dropna(subset=[wanted_columns.columns[0]])
        logger.debug(f"Extracted and cleaned DataFrame shape: {df_cleaned.shape}")

        result_dict = {}

        for _, row in df_cleaned.iterrows():
            if pd.notna(row[2]):
                result_dict[row.iloc[0]] = (row.iloc[1], row.iloc[2])
                logger.debug(f"Extracted entry: {row.iloc[0]} -> ({row.iloc[1]}, {row.iloc[2]})")
            else:
                result_dict[row.iloc[0]] = row.iloc[1]
                logger.debug(f"Extracted entry: {row.iloc[0]} -> {row.iloc[1]}")

        logger.info(f"Total inputs extracted from '{input_header}': {len(result_dict)}")
        return result_dict

    except Exception as e:
        logger.exception(f"Error in 'find_inputs_and_extract' for header '{input_header}': {e}")
        return {}

def find_membrane_correction_and_extract(doc_name, sheet_name, input_header):
    try:
        logger.debug(f"Loading workbook '{doc_name}'.")
        wb = openpyxl.load_workbook(doc_name)
        sheet = wb[sheet_name]
        logger.debug(f"Loaded sheet '{sheet_name}' from workbook '{doc_name}'.")

        column_value = None
        row_value = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == input_header:
                    column_value = cell.column
                    row_value = cell.row
                    logger.debug(f"Found input header '{input_header}' at column {column_value}, row {row_value}.")
                    break
            if column_value and row_value:
                break

        if not column_value or not row_value:
            logger.error(f"Input header '{input_header}' not found in sheet '{sheet_name}'.")
            return {}

        df = pd.read_excel(doc_name, sheet_name=sheet_name, header=row_value-1)
        wanted_columns = df.iloc[:, column_value-1:column_value +3]
        logger.debug(f"Selected columns {column_value-1} to {column_value +2}. Shape: {wanted_columns.shape}")

        AS_row_value = None
        for i, row in wanted_columns.iterrows():
            for j, cell in row.items():
                if cell == 'Axial Strain':
                    AS_row_value = i
                    logger.debug(f"Found 'Axial Strain' at row index {i}.")
                    break
            if AS_row_value is not None:
                break

        if AS_row_value is None:
            logger.error("'Axial Strain' not found in the selected columns.")
            return {}

        AS_and_more = wanted_columns.iloc[AS_row_value:, 0:3].dropna(subset=[wanted_columns.columns[0]])
        AS_and_more.columns = AS_and_more.iloc[0]
        AS_and_more = AS_and_more.iloc[1:]
        logger.debug(f"AS_and_more DataFrame shape after processing: {AS_and_more.shape}")

        result_dict = {col: AS_and_more[col].tolist() for col in AS_and_more.columns}
        logger.debug(f"Extracted membrane correction data: {result_dict}")

        actual_column_value = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 'Actual Diameter':
                    actual_column_value = cell.column
                    row_value = cell.row
                    logger.debug(f"Found 'Actual Diameter' at column {actual_column_value}, row {row_value}.")
                    break
            if actual_column_value and row_value:
                break

        if not actual_column_value or not row_value:
            logger.error("'Actual Diameter' not found in sheet '{sheet_name}'.")
            return {}

        for i, row in df.iterrows():
            if row.get('Actual Diameter') == 'Actual Diameter':
                actual_row_value = i
                logger.debug(f"Found 'Actual Diameter' at row index {i}.")
                break
        else:
            logger.error("'Actual Diameter' marker not found in DataFrame.")
            return {}

        actual_columns = df.iloc[actual_row_value:actual_row_value+3, actual_column_value-1:actual_column_value +2]
        logger.debug(f"Selected actual columns. Shape: {actual_columns.shape}")

        for _, row in actual_columns.iterrows():
            if pd.notna(row[2]):
                result_dict[row.iloc[0]] = (row.iloc[1], row.iloc[2])
                logger.debug(f"Extracted actual column entry: {row.iloc[0]} -> ({row.iloc[1]}, {row.iloc[2]})")
            else:
                result_dict[row.iloc[0]] = row.iloc[1]
                logger.debug(f"Extracted actual column entry: {row.iloc[0]} -> {row.iloc[1]}")

        # Find 'kPa/strain' column
        column_value = None
        row_value = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 'kPa/strain':
                    column_value = cell.column
                    row_value = cell.row
                    logger.debug(f"Found 'kPa/strain' at column {column_value}, row {row_value}.")
                    break
            if column_value and row_value:
                break

        if not column_value or not row_value:
            logger.error("'kPa/strain' not found in sheet '{sheet_name}'.")
            return {}

        x = pd.read_excel(doc_name, sheet_name=sheet_name)
        kpa_value = x.iloc[row_value-2, column_value]
        result_dict['kPa/strain'] = kpa_value
        logger.debug(f"Extracted 'kPa/strain' value: {kpa_value}")

        logger.info(f"Membrane correction extraction successful. Total entries: {len(result_dict)}")
        return result_dict

    except Exception as e:
        logger.exception(f"Error in 'find_membrane_correction_and_extract' for header '{input_header}': {e}")
        return {}

