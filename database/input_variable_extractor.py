import openpyxl
import pandas as pd

def find_inputs_and_extract(doc_name, sheet_name, input_header):
    wb = openpyxl.load_workbook(doc_name)
    sheet = wb[sheet_name]
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == input_header:
                column_value = cell.column
                row_value = cell.row
                break
    
    df = pd.read_excel(doc_name, sheet_name = sheet_name, header= row_value-1)
    wanted_columns = df.iloc[:, [column_value - 1, column_value, column_value + 1]]
    df_cleaned = wanted_columns.dropna(subset=[wanted_columns.columns[0]])

    result_dict = {}

    for _, row in df_cleaned.iterrows():
        if pd.notna(row[2]):
            result_dict[row.iloc[0]] = (row.iloc[1], row.iloc[2]) 
        else:
            result_dict[row.iloc[0]] = row.iloc[1]
    
    return result_dict
    
#print(find_inputs_and_extract('Copy of CSL_1_U-template.xlsx', '01 - Inputs', 'General inputs and calculations'))
#print(find_inputs_and_extract('Copy of CSL_1_U-template.xlsx', '01 - Inputs', 'Sample dimensions'))

def find_membrane_correction_and_extract(doc_name, sheet_name, input_header):
    wb = openpyxl.load_workbook(doc_name)
    sheet = wb[sheet_name]
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == input_header:
                column_value = cell.column
                row_value = cell.row
                break
            
    df = pd.read_excel(doc_name, sheet_name = sheet_name, header= row_value-1)
    wanted_columns = df.iloc[:, column_value-1:column_value +3]

    
    for i, row in wanted_columns.iterrows(): #Finding 'Axial Strain', 'St Correction' and 'Correction for this test'
        for j, cell in row.items(): 
            if cell == 'Axial Strain':
                AS_row_value = i
                break
    
    AS_and_more = wanted_columns.iloc[AS_row_value:, 0:3].dropna(subset=[wanted_columns.columns[0]])
    AS_and_more.columns = AS_and_more.iloc[0]
    AS_and_more = AS_and_more.iloc[1:]

    result_dict = {col: AS_and_more[col].tolist() for col in AS_and_more.columns}
    
    for row in sheet.iter_rows(): #Finding column index of 'Actual Diameter', 'Actual Membrane Thickness' and 'Mutiple Correction by'
        for cell in row:
            if cell.value == 'Actual Diameter':
                actual_column_value = cell.column
                break
    
    for i, row in df.iterrows(): #Finding row index
        for j, cell in row.items(): 
            if cell == 'Actual Diameter':
                actual_row_value = i
                break
    
    actual_columns = df.iloc[actual_row_value:actual_row_value+3, actual_column_value-1: actual_column_value +2]
    
    for _, row in actual_columns.iterrows():
        if pd.notna(row[2]):
            result_dict[row.iloc[0]] = (row.iloc[1], row.iloc[2]) 
        else:
            result_dict[row.iloc[0]] = row.iloc[1]
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'kPa/strain':
                column_value = cell.column
                row_value = cell.row
                print(column_value, row_value)
                break

    x = pd.read_excel(doc_name, sheet_name = sheet_name)

    kpa_value = x.iloc[row_value-2, column_value]
    
    result_dict['kPa/strain']= x.iloc[row_value-2, column_value]
    
    return result_dict

#print(find_membrane_correction_and_extract('Copy of CSL_1_U-template.xlsx', '01 - Inputs', 'Membrane correction'))

